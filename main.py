"""Download stock data from the Internet, format it, and make a CSV."""
import csv
import json
from pyperclip import copy
import socket
import sys
import time
from tkinter import Tk
from base64 import b64decode
from datetime import datetime
from os.path import expanduser

import openpyxl
import requests

__version__ = '1.1.6'

REPO = 'http://api.github.com/repos/carter-lavering/Alfred/'


def api(subdir):
    """Return the GitHub API response to subdir."""
    return requests.get(REPO + subdir)


def get_latest_release():
    """Return whether this script is outdated."""
    response = api('releases')
    releases = [x['tag_name'][1:] for x in response.json()]
    return sorted(releases)[-1]


def replace_with_latest():
    """Download the latest version and replace this file with it."""
    response = api('contents/main.py')
    encoded = response.json()['content']
    decoded = str(b64decode(encoded), 'utf-8')
    with open(__file__, 'w') as file:
        file.write(decoded)


def self_update():
    """Update only if outdated."""
    print('Checking for updates...', end=' ', flush=True)
    latest_version = get_latest_release()
    if latest_version > __version__:
        print('Update found')
        print(
            'Updating to {new} from {old}...'.format(
                new=latest_version, old=__version__),
            end=' ',
            flush=True)
        replace_with_latest()
        print('Done')
    else:
        print('No update found')


# \_\_\_\_    \_\_\_\_\_  \_\_\_\_\_  \_\_\_\_\_  \_      \_  \_\_\_\_
#  \_      \_  \_          \_              \_      \_\_    \_  \_
#   \_      \_  \_\_\_\_\_  \_\_\_\_\_      \_      \_  \_  \_  \_\_\_\_
#    \_      \_  \_          \_              \_      \_    \_\_  \_
#     \_\_\_\_    \_\_\_\_\_  \_          \_\_\_\_\_  \_      \_  \_\_\_\_


def get_sheet_corner(workbook_path, sheet_name=None):
    """Return the column and row of the upper left corner of a spreadsheet.

    Indexing starts at 1, so A1 is (1, 1), not (0, 0).
    """
    # I have to use x and y because rows and columns get me confused about
    # which way they go
    wb = openpyxl.load_workbook(workbook_path)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    first_x = 0
    corner_found = False
    while not corner_found:
        if first_x >= 1000:
            raise RuntimeError('No data found for 1000 columns')
        for x in range(first_x, -1, -1):
            y = first_x - x
            temp_cell = ws.cell(row=y + 1, column=x + 1)
            if temp_cell.value:
                return (x + 1, y + 1)
                corner_found = True
        first_x += 1


def copy(text):
    """Copies text to clipboard."""
    r = Tk()
    r.withdraw()
    r.clipboard_clear()
    r.clipboard_append(repr(text))

    r.update()
    time.sleep(.2)
    r.update()

    # r.destroy()


def read_sheet_column(workbook_path,
                      sheet_name=None,
                      headers=True,
                      select=False):
    """Read the first column in a given sheet.

    If headers is True, then loop through all the cells below the upper-left
    corner until a blank space is found. Return a list of all the cells. If a
    cell has a hashtag in the cell to the left of it, do not return that cell.
    """
    corner = get_sheet_corner(workbook_path, sheet_name)
    wb = openpyxl.load_workbook(workbook_path)
    output_cells = []
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    x = corner[0]
    if headers:
        y = corner[1] + 1  # Don't want the headers in the data
    else:
        y = corner[1]
    read_cell = ws.cell(row=y, column=x)
    while read_cell.value:
        read_cell = ws.cell(row=y, column=x)
        if x == 1:
            output_cells.append(read_cell.value)
        else:
            adjacent_cell = ws.cell(row=y, column=x - 1)
            if select:
                if '#' in str(adjacent_cell.value):
                    try:
                        output_cells.append(read_cell.value.upper())
                    except AttributeError:
                        output_cells.append(read_cell.value)
            else:
                if '#' not in str(adjacent_cell.value):
                    try:
                        output_cells.append(read_cell.value.upper())
                    except AttributeError:
                        output_cells.append(read_cell.value)
        y += 1
    return [cell for cell in output_cells if cell is not None]


def strip_matrix(matrix):
    """Ensure matrix is shifted as far left as possible."""
    while not any([row[0] for row in matrix]):
        matrix = [row[1:] for row in matrix]
    return matrix


def read_sheet(workbook_path, sheet_name=None, headers=True):
    """Return 2D list of cell values from uncommented rows."""
    workbook = openpyxl.load_workbook(workbook_path)
    if sheet_name is not None:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active
    table = strip_matrix([[cell.value for cell in row] for row in sheet])
    uncommented_rows = [
        row[1:] for row in table
        if row[0] != '#' if any([cell is not None for cell in row])
    ]
    return uncommented_rows


def week(timestamp):
    """Return the ISO calendar week number of a given timestamp.

    Timestamp can be either an integer or a string.
    """
    return datetime.utcfromtimestamp(int(timestamp)).isocalendar()[1]


def end_script(terminate=True):
    """End program."""
    if not isdev:
        input('Press enter to exit')
        sys.exit()
    elif terminate:
        sys.exit()


def rearrange(lst, order):
    """Return lst but in the order of order.

    Indexing starts at 0.
    """
    return [lst[x] for x in order]


def mass_lookup(d, k):
    """Return a list of the values of keys k from d, ignoring errors."""
    output = []
    for key in k:
        try:
            output.append(d[key])
        except KeyError:
            print('Key {k} not found in {d}'.format(k=key, d=d))
    return output


# \_\_\_\_\_  \_\_\_\_\_  \_\_\_\_\_  \_\_\_\_    \_\_\_\_\_
#  \_              \_      \_      \_  \_      \_      \_
#   \_\_\_\_\_      \_      \_\_\_\_\_  \_\_\_\_        \_
#            \_      \_      \_      \_  \_    \_        \_
#     \_\_\_\_\_      \_      \_      \_  \_      \_      \_


def main():
    print('Alfred version {0}'.format(__version__))
    isdev = socket.gethostname() == 'raphael'

    desktop = expanduser('~') + '\\Desktop\\'

    print('Opening files...')
    try:
        symbols_sheet = read_sheet(desktop + 'stock_signs.xlsx')
        signs = [row[0] for row in symbols_sheet]
        assert len(symbols_sheet) == len(signs)
        additional_data = {}
        for row in symbols_sheet:
            additional_data[row[0]] = dict()
            additional_data[row[0]]['Ex dividend date'] = row[4]
            additional_data[row[0]]['Quarterly dividend'] = row[5]
            additional_data[row[0]]['Capitalization'] = row[6]
            additional_data[row[0]]['Next earnings date'] = row[7]
            additional_data[row[0]]['Rating'] = row[8]
    except FileNotFoundError:
        write_signs = openpyxl.Workbook()
        write_signs.save(desktop + 'stock_signs.xlsx')
        print('Please go to your desktop and put the symbols you want into'
              ' stock_signs.xlsx. Put hash marks in the cells to the left of'
              ' the ones you don\'t want.')
        end_script(terminate=False)
    try:
        dates_sheet = read_sheet(desktop + 'stock_dates.xlsx')
        dates = [row[0] for row in dates_sheet]
        dates_weeks = [date.isocalendar()[:1] for date in dates]
    except FileNotFoundError:
        write_dates = openpyxl.Workbook()
        write_dates.save(desktop + 'stock_dates.xlsx')
        print('Please go to your desktop and put the dates you want into'
              ' stock_dates.xlsx. Put hash marks in the cells to the left of'
              " the ones you don't want.")
        end_script(terminate=False)

    assert signs
    assert dates

    print('{0} signs, {1} dates'.format(len(signs), len(dates)))

    dt = datetime.fromtimestamp(time.time())
    date = dt.strftime('%d-%m-%Y')

    if not isdev:
        output_path = (
            'C:/Users/Gary/Documents/Option_tables/Option_Model_Files/'
            'OptionReportDirectory/options_report_{0}.csv'.format(date))
    else:
        output_path = 'options_report_{0}.csv'.format(date)

    output_name = output_path.split('/')[-1]

    start = time.time()

    # \_\_\_\_      \_\_\_    \_      \_  \_      \_
    #  \_      \_  \_      \_  \_      \_  \_\_    \_
    #   \_      \_  \_      \_  \_  \_  \_  \_  \_  \_  \_\_\_\_\_
    #    \_      \_  \_      \_  \_  \_  \_  \_    \_\_
    #     \_\_\_\_      \_\_\_      \_  \_    \_      \_

    #       \_            \_\_\_      \_\_\_    \_\_\_\_
    #        \_          \_      \_  \_      \_  \_      \_
    #         \_          \_      \_  \_\_\_\_\_  \_      \_
    #          \_          \_      \_  \_      \_  \_      \_
    #           \_\_\_\_\_    \_\_\_    \_      \_  \_\_\_\_

    options_data_url = 'https://query1.finance.yahoo.com/v7/finance/options/{0}'
    stock_data_url = ('https://query1.finance.yahoo.com/v10/finance/quoteSummary/'
                      '{0}?modules=assetProfile')

    # Headers
    all_data = [[
        'Stock', 'Timestamp', 'Contract Symbol', 'Strike', 'Currency',
        'Last Price', 'Change', '% Change', 'Volume', 'Open Interest', 'Bid',
        'Ask', 'Contract Size', 'Expiration', 'Last Trade Date',
        'Implied Volatility', 'In The Money', 'Stock Last', 'Industry', 'Sector',
        'Company'
    ]]
    json_headers = [
        'contractSymbol', 'strike', 'currency', 'lastPrice', 'change',
        'percentChange', 'volume', 'openInterest', 'bid', 'ask', 'contractSize',
        'expiration', 'lastTradeDate', 'impliedVolatility', 'inTheMoney',
        'quoteLast', 'industry', 'sector', 'company'
    ]
    errors = []

    first_iter = True
    for sign in signs:
        # Print "SYMBOL (2 of 4)"
        print(
            '{n}{0:{1}} ({2:{3}} of {4})'.format(
                sign,
                len(max(signs, key=len)),
                signs.index(sign) + 1,
                len(str(len(signs))),
                len(signs),
                n='' if first_iter else '\n'),
            end='')

        dates_page = requests.get(options_data_url.format(sign))
        dates_json = json.loads(dates_page.text)
        try:
            timestamps_from_site = (
                dates_json['optionChain']['result'][0]['expirationDates'])
        except (IndexError, TypeError) as e:
            print(' Non-existent', end='')
            continue

        # timestamps_to_use = timestamps_from_site
        timestamps_to_use = [
            ts for ts in timestamps_from_site
            if datetime.fromtimestamp(ts).isocalendar()[:1] in dates_weeks
        ]
        print(' [', '-' * len(timestamps_to_use), ']', sep='', end='', flush=True)

        weekdays = []

        stock_page = requests.get(stock_data_url.format(sign))
        stock_json = json.loads(stock_page.text)

        profile = stock_json['quoteSummary']['result'][0]['assetProfile']
        try:
            industry, sector = profile['industry'], profile['sector']
        except KeyError:
            print(
                ' Sector and industry unavailable',
                '\b' * 32,
                sep='',
                end='',
                flush=True)
            industry = sector = ''

        print('\b' * (len(timestamps_to_use) + 1), end='')

        messages = []
        for ts in timestamps_to_use:
            complete_success = True
            try:
                data_page = requests.get(
                    options_data_url.format(sign) + '?date=' + str(ts))
            # Try twice more
            except TimeoutError:
                try:
                    data_page = requests.get(
                        options_data_url.format(sign) + '?date=' + str(ts))
                except TimeoutError:
                    # TODO: More verbose
                    print('-', end='', flush=True)
                    messages.append('{d} timed out'.format(
                        d=datetime.utcfromtimestamp(ts).strftime('%m/%d/%Y')))
                    continue
            try:
                data_json = json.loads(data_page.text)
            except ValueError:
                complete_success = False
                messages.append('Can\'t decode json')

            specific_data = data_json['optionChain']['result'][0]
            # {'Stock Last': specific_data['quote']['regularMarketPrice']}
            data_dict = (specific_data['options'][0]['calls'])  # List of dicts
            for row in data_dict:
                row.update({
                    'quoteLast':
                    specific_data['quote']['regularMarketPrice'],
                    'company':
                    specific_data['quote']['longName'],
                    'industry':
                    industry,
                    'sector':
                    sector
                })
                try:
                    all_data.append([sign, start] +
                                    [row[key] for key in json_headers])
                except KeyError:
                    # TODO: More verbose
                    complete_success = False
                    messages.append('Something went wrong with {d} ({ts})'.format(
                        d=datetime.utcfromtimestamp(ts).strftime('%m/%d/%Y'),
                        ts=ts))
                    print('-', end='', flush=True)
                    continue
            if complete_success:
                print('=', end='', flush=True)
        if messages:
            print('] ', ', '.join(messages), end='', flush=True)
        first_iter = False

    print()  # Allow printing of the last line

    # \_\_\_\_\_    \_\_\_    \_\_\_\_    \_      \_  \_\_\_\_\_  \_\_\_\_\_
    #  \_          \_      \_  \_      \_  \_\_  \_\_  \_      \_      \_
    #   \_\_\_\_    \_      \_  \_\_\_\_    \_  \_  \_  \_\_\_\_\_      \_
    #    \_          \_      \_  \_    \_    \_      \_  \_      \_      \_
    #     \_            \_\_\_    \_      \_  \_      \_  \_      \_      \_

    headers = [
        'Symbol', 'Company', 'Industry', 'Sector', 'Ex dividend date',
        'Quarterly dividend', 'Capitalization', 'Rating', 'Next earnings date',
        'Price', 'Expiration', 'Strike', 'Bid', 'Ask', 'Volume', 'Last Call',
        datetime.now().date(), 'days', '70,000', ' $invested', '$prem', 'prem%',
        'annPrem%', 'MaxRet', 'Max%', 'annMax%', '10%'
    ]

    all_data_by_header = [{h: x[i]
                           for i, h in enumerate(all_data[0])}
                          for x in all_data[1:]]

    for d in all_data_by_header:
        d['Timestamp'] = (
            datetime.utcfromtimestamp(d['Timestamp']).strftime('%m/%d/%Y %H:%M'))

        d['Last Trade Date'] = (
            datetime.utcfromtimestamp(d['Last Trade Date']).strftime('%m/%d/%Y'))

        d['Expiration'] = (
            datetime.utcfromtimestamp(d['Expiration']).strftime('%m/%d/%Y'))

        d.update(additional_data[d['Stock']])

    formulas = [
        '=IF(P{n}<N{n},(P{n}-N{n})+T{n},T{n})', '=O{n}-U$6',
        '=ROUND(W$6/((N{n}-0)*100),0)', '=100*W{n}*(N{n}-0)', '=100*U{n}*W{n}',
        '=Y{n}/X{n}', '=(365/V{n})*Z{n}',
        '=IF(P{n}>N{n},(100*W{n}*(P{n}-N{n}))+Y{n},Y{n})', '=AB{n}/X{n}',
        '=(365/V{n})*AC{n}', '=IF((ABS(P{n}-N{n})/P{n})<AE$6,"NTM","")'
    ]

    v_offset = 5
    h_offset = 4

    keys = [
        'Stock', 'Company', 'Industry', 'Sector', 'Ex dividend date',
        'Quarterly dividend', 'Capitalization', 'Rating', 'Next earnings date',
        'Stock Last', 'Expiration', 'Strike', 'Bid', 'Ask', 'Volume', 'Last Price'
    ]
    # +2 because Excel starts counting at 1 and because there are headers
    formatted_data_table = ([headers] + [
        mass_lookup(row, keys) + [f.format(n=i + v_offset + 2) for f in formulas]
        for i, row in enumerate(all_data_by_header)
    ])

    # Offset for formulas to work
    formatted_data_table = ([[]] * v_offset + [[''] * h_offset + row
                                               for row in formatted_data_table])

    # # \_      \_  \_\_\_\_    \_\_\_\_\_  \_\_\_\_\_  \_\_\_\_\_
    # #  \_      \_  \_      \_      \_          \_      \_
    # #   \_      \_  \_\_\_\_        \_          \_      \_\_\_\_\_
    # #    \_  \_  \_  \_    \_        \_          \_      \_
    # #     \_\_  \_\_  \_      \_  \_\_\_\_\_      \_      \_\_\_\_\_

    print('Writing to {0}...'.format(output_name), end=' ', flush=True)

    try:
        with open(output_path, 'w', newline='') as csv_file:
            csv_writer = csv.writer(csv_file)
            for row in formatted_data_table:
                csv_writer.writerow(row)
    except PermissionError:
        print('\a\nPlease close {0}.'.format(output_name))
        input('Press enter when done')
        print('Writing to {0}...'.format(output_name, end=' ', flush=True))
        with open(output_path, 'w', newline='') as csv_file:
            csv_writer = csv.writer(csv_file)
            for row in formatted_data_table:
                csv_writer.writerow(row)
    except FileNotFoundError:
        try:
            with open('options_report_{date}.csv'.format(date=date), 'w', newline='') as csv_file:
                csv_writer = csv.writer(csv_file)
                for row in formatted_data_table:
                    csv_writer.writerow(row)
        except PermissionError:
            pass

    print('Done')


if __name__ == '__main__':
    try:
        self_update()
        main()
    except Exception as e:
        print("Something's gone wrong. Here's the error:")
        print()
        print(repr(e))
        print()
        input('Please select the error with your mouse, right-click to copy, and paste with Ctrl-V into an email.\nPress Enter to exit')
